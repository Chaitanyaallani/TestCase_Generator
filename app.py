import streamlit as st
from groq import Groq
import pytesseract
from PIL import Image, ImageEnhance
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sentence_transformers import SentenceTransformer
import io, os, re, json, math
import zipfile
from xml.etree import ElementTree as ET

# ── Page Config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="QA Test Case Generator",
    page_icon="🧪",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;700&family=IBM+Plex+Sans:wght@300;400;500;600&display=swap');
*, html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
.stApp { background: #F7F6F3; }
.app-header {
    background: #1A1A2E; border-radius: 0 0 24px 24px;
    padding: 2.5rem 2rem 2rem; margin: -1rem -1rem 2rem -1rem;
    position: relative; overflow: hidden;
}
.app-header::after {
    content: '🧪'; position: absolute; right: 2rem; top: 50%;
    transform: translateY(-50%); font-size: 5rem; opacity: 0.08;
}
.app-title { font-family: 'IBM Plex Mono', monospace; font-size: 2rem; font-weight: 700; color: #E8F4FD; margin: 0; }
.app-sub   { color: #7B8FA1; font-size: 0.9rem; margin-top: 0.4rem; }
.tag {
    display: inline-block; background: rgba(255,255,255,0.07);
    border: 1px solid rgba(255,255,255,0.12); color: #A8C8E8;
    padding: 0.15rem 0.65rem; border-radius: 4px; font-size: 0.7rem;
    font-family: 'IBM Plex Mono', monospace; margin: 0.3rem 0.2rem 0 0;
}
.card-label { font-family: 'IBM Plex Mono', monospace; font-size: 0.65rem; color: #9B8EA0; text-transform: uppercase; letter-spacing: 2px; margin-bottom: 0.5rem; }
.result-preview { background: #1A1A2E; border-radius: 8px; padding: 1rem 1.2rem; font-family: 'IBM Plex Mono', monospace; font-size: 0.75rem; color: #A8C8E8; white-space: pre-wrap; max-height: 300px; overflow-y: auto; margin-top: 0.5rem; }
.metric-strip { display: flex; gap: 1rem; margin: 1rem 0; }
.metric { flex: 1; background: #FFFFFF; border: 1px solid #E8E5DF; border-radius: 10px; padding: 1rem; text-align: center; }
.metric-val { font-family: 'IBM Plex Mono', monospace; font-size: 1.8rem; font-weight: 700; color: #1A1A2E; }
.metric-lbl { font-size: 0.7rem; color: #9B8EA0; text-transform: uppercase; letter-spacing: 1px; margin-top: 0.2rem; }
.success-box { background: #E8F5E9; border: 1px solid #A5D6A7; border-left: 4px solid #2E7D32; border-radius: 8px; padding: 1.2rem 1.5rem; margin: 1rem 0; color: #1B5E20; font-weight: 500; }
.warn-box  { background: #FFF3CD; border: 1px solid #FFD580; border-left: 4px solid #F59E0B; border-radius: 8px; padding: 0.8rem 1.2rem; margin: 0.5rem 0; color: #92400E; font-size: 0.88rem; }
.info-box  { background: #E3F2FD; border: 1px solid #90CAF9; border-left: 4px solid #1565C0; border-radius: 8px; padding: 0.8rem 1.2rem; margin: 0.5rem 0; color: #0D47A1; font-size: 0.88rem; }
.rag-ok   { background: #E8F5E9; border: 1px solid #A5D6A7; border-left: 3px solid #2E7D32; border-radius: 6px; padding: 0.6rem 0.8rem; font-size: 0.8rem; color: #1B5E20; margin-top: 0.5rem; }
.rag-warn { background: #FFF3CD; border: 1px solid #FFD580; border-left: 3px solid #F59E0B; border-radius: 6px; padding: 0.6rem 0.8rem; font-size: 0.8rem; color: #92400E; margin-top: 0.5rem; }
.llm-badge { display: inline-block; background: #E3F2FD; border: 1px solid #90CAF9; color: #0D47A1; padding: 0.2rem 0.6rem; border-radius: 4px; font-size: 0.75rem; font-family: 'IBM Plex Mono', monospace; }
.stButton > button { background: #1A1A2E !important; color: #E8F4FD !important; border: none !important; border-radius: 8px !important; font-family: 'IBM Plex Mono', monospace !important; font-size: 0.82rem !important; font-weight: 600 !important; letter-spacing: 1px !important; padding: 0.65rem 1.5rem !important; width: 100% !important; transition: all 0.2s !important; }
.stButton > button:hover { background: #2D2D4E !important; transform: translateY(-1px) !important; box-shadow: 0 4px 16px rgba(26,26,46,0.25) !important; }
.stButton > button:disabled { background: #C8C5BF !important; color: #8A8580 !important; }
[data-testid="stSidebar"] { background: #FFFFFF; border-right: 1px solid #E8E5DF; }
.sidebar-section { background: #F7F6F3; border: 1px solid #E8E5DF; border-radius: 10px; padding: 1rem; margin-bottom: 1rem; }
.sidebar-title { font-family: 'IBM Plex Mono', monospace; font-size: 0.65rem; text-transform: uppercase; letter-spacing: 2px; color: #9B8EA0; margin-bottom: 0.6rem; }
#MainMenu, footer, header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# LLM CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

LLM_PROVIDERS = {
    "Groq — LLaMA 3.1 8B (fastest)": {
        "provider" : "groq",
        "model"    : "llama-3.1-8b-instant",
        "key_hint" : "gsk_...",
        "key_link" : "https://console.groq.com",
        "limit"    : "14,400 req/day free",
    },
    "Groq — LLaMA 3.3 70B (smarter)": {
        "provider" : "groq",
        "model"    : "llama-3.3-70b-versatile",
        "key_hint" : "gsk_...",
        "key_link" : "https://console.groq.com",
        "limit"    : "1,000 req/day free",
    },
    "Groq — Gemma 2 9B": {
        "provider" : "groq",
        "model"    : "gemma2-9b-it",
        "key_hint" : "gsk_...",
        "key_link" : "https://console.groq.com",
        "limit"    : "14,400 req/day free",
    },
    "Groq — Mixtral 8x7B": {
        "provider" : "groq",
        "model"    : "mixtral-8x7b-32768",
        "key_hint" : "gsk_...",
        "key_link" : "https://console.groq.com",
        "limit"    : "14,400 req/day free",
    },
    "Together AI — LLaMA 3 8B (free tier)": {
        "provider" : "together",
        "model"    : "meta-llama/Llama-3-8b-chat-hf",
        "key_hint" : "...",
        "key_link" : "https://api.together.xyz",
        "limit"    : "$25 free credits",
    },
}


def call_llm(prompt: str, api_key: str, provider_config: dict, max_tokens: int = 8000) -> str:
    """Call whichever LLM provider is selected."""
    provider = provider_config["provider"]
    model    = provider_config["model"]

    if provider == "groq":
        client   = Groq(api_key=api_key)
        response = client.chat.completions.create(
            model      = model,
            messages   = [{"role": "user", "content": prompt}],
            max_tokens = max_tokens,
        )
        return response.choices[0].message.content

    elif provider == "together":
        import urllib.request as ur
        body = json.dumps({
            "model"      : model,
            "messages"   : [{"role": "user", "content": prompt}],
            "max_tokens" : max_tokens,
        }).encode()
        req = ur.Request(
            "https://api.together.xyz/v1/chat/completions",
            data    = body,
            headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        )
        with ur.urlopen(req, timeout=120) as resp:
            data = json.loads(resp.read())
        return data["choices"][0]["message"]["content"]

    raise ValueError(f"Unknown provider: {provider}")


# ══════════════════════════════════════════════════════════════════════════════
# HELPER — SAFE EXCEL SHEET NAME
# Fixes ValueError from openpyxl when feature name contains illegal characters.
# Excel sheet names cannot contain: : \ / ? * [ ]
# Sheet names also cannot be empty, start/end with apostrophe, or exceed 31 chars.
# ══════════════════════════════════════════════════════════════════════════════
def safe_sheet_name(name: str, fallback: str = "TestCases") -> str:
    """
    Sanitise a string so it is a valid Excel worksheet name.
    - Removes/replaces illegal characters: : \\ / ? * [ ]
    - Strips leading/trailing spaces and apostrophes
    - Truncates to 31 characters (Excel hard limit)
    - Returns fallback if result is empty
    """
    if not name or not name.strip():
        return fallback

    # Replace every illegal Excel sheet-name character with underscore
    sanitised = re.sub(r'[:\\/?*\[\]]', '_', name)

    # Remove any leading/trailing whitespace and apostrophes
    sanitised = sanitised.strip(" '")

    # Collapse multiple consecutive underscores into one
    sanitised = re.sub(r'_+', '_', sanitised)

    # Truncate to Excel's 31-character sheet-name limit
    sanitised = sanitised[:31].strip(" '_")

    return sanitised if sanitised else fallback


# ══════════════════════════════════════════════════════════════════════════════
# IN-MEMORY VECTOR STORE
# ══════════════════════════════════════════════════════════════════════════════

class SimpleVectorStore:
    def __init__(self):
        self.documents  : list  = []
        self.embeddings : list  = []
        self.ids        : set   = set()

    def add(self, documents: list, embeddings: list, ids: list):
        for doc, emb, id_ in zip(documents, embeddings, ids):
            if id_ not in self.ids:
                self.documents.append(doc)
                self.embeddings.append(emb)
                self.ids.add(id_)

    def count(self) -> int:
        return len(self.documents)

    def query(self, query_embeddings: list, n_results: int = 10) -> dict:
        if not self.documents:
            return {"documents": [[]]}
        qe = query_embeddings[0]
        def cosine(a: list, b: list) -> float:
            dot = sum(x * y for x, y in zip(a, b))
            na  = math.sqrt(sum(x * x for x in a))
            nb  = math.sqrt(sum(x * x for x in b))
            return dot / (na * nb) if na and nb else 0.0
        scored = [(cosine(qe, e), d) for e, d in zip(self.embeddings, self.documents)]
        scored.sort(key=lambda x: x[0], reverse=True)
        return {"documents": [[d for _, d in scored[:n_results]]]}

    def clear(self):
        self.documents  = []
        self.embeddings = []
        self.ids        = set()


@st.cache_resource(show_spinner="Loading embedding model...")
def load_embed_model():
    return SentenceTransformer("paraphrase-MiniLM-L3-v2")

@st.cache_resource
def get_vector_store() -> SimpleVectorStore:
    return SimpleVectorStore()


# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
DEFAULTS = {
    "extracted_text"    : None,
    "feature_understand": None,
    "test_cases_parsed" : [],
    "excel_bytes"       : None,
    "stage"             : 0,
    "tc_count"          : 0,
    "pos_count"         : 0,
    "neg_count"         : 0,
    "edge_count"        : 0,
    "rag_loaded"        : False,
    "rag_count"         : 0,
    "rag_docs"          : [],
    "last_tc_id"        : 3000,
    "ocr_warnings"      : [],
    "llm_used"          : "",
}
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ══════════════════════════════════════════════════════════════════════════════
# OCR
# ══════════════════════════════════════════════════════════════════════════════
def run_ocr(image: Image.Image) -> tuple:
    """Extract text from image with preprocessing. Returns (text, warnings)."""
    def try_ocr(img, config):
        try:
            text = pytesseract.image_to_string(img, config=config).strip()
            return text if len(text) > 30 else ""
        except Exception:
            return ""

    img = image.convert("RGB")
    w, h = img.size
    if w < 1000:
        img = img.resize((int(w * 1000 / w), int(h * 1000 / w)), Image.LANCZOS)
    img = ImageEnhance.Contrast(img).enhance(2.0)
    img = ImageEnhance.Sharpness(img).enhance(2.0)
    gray = img.convert("L")

    for config in ["--psm 6 --oem 3", "--psm 11", "--psm 3"]:
        text = try_ocr(gray, config)
        if text:
            return text, []

    text = try_ocr(image.convert("RGB"), "--psm 6")
    if text:
        return text, []

    return "", ["⚠️ OCR could not extract text from this image. Please paste the feature description in the Text tab."]


def extract_docx_text(file_bytes: bytes) -> str:
    """Extract text from .docx files using XML parsing."""
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            if "word/document.xml" not in z.namelist():
                return ""
            xml = z.read("word/document.xml")
        NS   = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
        tree = ET.fromstring(xml)
        paras = []
        for para in tree.iter(f"{{{NS}}}p"):
            text = "".join(n.text or "" for n in para.iter(f"{{{NS}}}t"))
            if text.strip():
                paras.append(text.strip())
        return "\n".join(paras)
    except Exception:
        return ""


def extract_txt_text(file_bytes: bytes) -> str:
    for enc in ["utf-8", "latin-1", "cp1252"]:
        try:
            return file_bytes.decode(enc).strip()
        except Exception:
            pass
    return ""


# ══════════════════════════════════════════════════════════════════════════════
# RAG
# ══════════════════════════════════════════════════════════════════════════════
def load_excel_to_rag(excel_bytes: bytes) -> tuple:
    embed = load_embed_model()
    store = get_vector_store()
    total, last_id = 0, 3000
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    for sname in wb.sheetnames:
        ws = wb[sname]
        if ws.max_row < 3:
            continue
        headers = [
            str(c.value).strip().lower() if c.value else f"col{i}"
            for i, c in enumerate(ws[1])
        ]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            for num in re.findall(r'\d+', str(row[0] or "")):
                last_id = max(last_id, int(num))
            row_text = "\n".join(
                f"{headers[i]}: {v}"
                for i, v in enumerate(row)
                if v is not None and i < len(headers)
            )
            if not row_text.strip():
                continue
            st.session_state.rag_docs.append(row_text)
            emb = embed.encode(row_text).tolist()
            try:
                store.add(
                    documents  = [row_text],
                    embeddings = [emb],
                    ids        = [f"tc_{sname}_{total}"]
                )
                total += 1
            except Exception:
                pass
    st.session_state.rag_count  = total
    st.session_state.rag_loaded = True
    st.session_state.last_tc_id = last_id
    return total, last_id


def rag_retrieve(query: str) -> str:
    embed = load_embed_model()
    store = get_vector_store()
    if store.count() == 0 and st.session_state.rag_docs:
        for i, doc in enumerate(st.session_state.rag_docs):
            emb = embed.encode(doc).tolist()
            try:
                store.add(documents=[doc], embeddings=[emb], ids=[f"tc_r_{i}"])
            except Exception:
                pass
    if store.count() == 0:
        return "No past test cases loaded."
    emb     = embed.encode(query).tolist()
    results = store.query(query_embeddings=[emb], n_results=min(10, store.count()))
    docs    = results["documents"][0] if results["documents"] else []
    return "\n\n---\n\n".join(docs) if docs else "No similar cases found."


# ══════════════════════════════════════════════════════════════════════════════
# JSON PARSER — 4 fallback strategies
# ══════════════════════════════════════════════════════════════════════════════
def parse_json_response(raw: str) -> list:
    strategies = [
        lambda s: json.loads(s),
        lambda s: json.loads(re.search(r'\[[\s\S]*\]', s).group()),
        lambda s: json.loads(re.sub(r'```json|```', '', s).strip()),
        lambda s: json.loads(re.search(r'\[[\s\S]*\]',
                    re.sub(r',\s*([}\]])', r'\1',
                    re.sub(r"'", '"', s))).group()),
    ]
    for strategy in strategies:
        try:
            result = strategy(raw)
            if isinstance(result, list) and len(result) > 0:
                return result
        except Exception:
            pass
    return []


# ══════════════════════════════════════════════════════════════════════════════
# TEST CASE GENERATION
# ══════════════════════════════════════════════════════════════════════════════
def generate_test_cases(
    combined_text      : str,
    feature_understand : str,
    similar_cases      : str,
    num_cases          : int,
    auto_mode          : bool,
    last_tc_id         : int,
    api_key            : str,
    provider_config    : dict,
    include_neg        : bool,
    include_edge       : bool,
) -> list:

    start_id = last_tc_id + 1

    if auto_mode:
        count_rule = (
            "Generate as many test cases as needed for COMPLETE coverage — "
            "cover every acceptance criterion, every module, every user flow, "
            "every error state, and all edge/boundary conditions. "
            "Do NOT stop at any fixed number. Typically 40-80 for multi-module features."
        )
        type_rule = "Balance: ~40% Positive, ~35% Negative, ~25% Edge cases."
    else:
        if include_neg and include_edge:
            pos  = num_cases // 3 + (num_cases % 3)
            neg  = num_cases // 3
            edge = num_cases - pos - neg
        elif include_neg:
            pos  = num_cases // 2 + (num_cases % 2)
            neg  = num_cases - pos
            edge = 0
        elif include_edge:
            pos  = num_cases // 2 + (num_cases % 2)
            neg  = 0
            edge = num_cases - pos
        else:
            pos, neg, edge = num_cases, 0, 0
        count_rule = f"Generate exactly {num_cases} test cases."
        type_rule  = f"{pos} Positive + {neg} Negative + {edge} Edge cases."

    prompt = f"""You are a Senior QA Architect generating enterprise UI test cases.

TASK: {count_rule}
Types: {type_rule}
TC IDs start from: TC-{start_id:04d}

FEATURE UNDERSTANDING:
{feature_understand}

PAST TEST CASES — MATCH THIS EXACT STYLE AND QUALITY:
{similar_cases}

QUALITY RULES (MANDATORY):
1. Prerequisites: "User: [specific role] logged in, [specific module] accessible, [specific test data] available"
2. Steps: Exactly 5 steps. Each step = ONE specific UI action on a named UI element.
   Format: "1. Navigate to [module]\\n2. Click [button name]\\n3. Enter [value] in [field name]\\n4. Click [button]\\n5. Verify [exact outcome]"
3. Expected Result: Describe EXACTLY what the user sees — screen name + message text + element state change.
4. FORBIDDEN in Expected Result: "works correctly", "saves successfully", "behaves as expected", "is correct"
5. Test Type: exactly Positive / Negative / Edge
6. Priority: High = core AC flows, Medium = validations, Low = boundary/cosmetic

FEATURE INPUT:
{combined_text[:1200]}

RETURN ONLY a valid JSON array. No explanation. No markdown. No text before [ or after ].

Example of required quality:
[
  {{
    "id": "TC-{start_id:04d}",
    "title": "Verify year format change in incident report template displays correctly",
    "prerequisites": "User: ignio admin logged in, Incident report template accessible, Test incident data with date Jan 23 2026 available",
    "steps": "1. Navigate to incident management system\\n2. Open existing incident with date Jan 23 2026\\n3. Generate incident report using standard template\\n4. Review the date format in the generated report\\n5. Verify year is displayed in the new required format",
    "expected": "Incident report displays date in correct year format. Template shows date as per new standard. No formatting errors visible in the report.",
    "type": "Positive",
    "priority": "High",
    "related_tc": "None"
  }},
  {{
    "id": "TC-{start_id+1:04d}",
    "title": "Verify system rejects invalid year format input with clear error message",
    "prerequisites": "User: ignio admin logged in, Template configuration screen accessible",
    "steps": "1. Navigate to template configuration screen\\n2. Locate the year format input field\\n3. Enter invalid year format value '202' (3 digits only)\\n4. Click the Save button\\n5. Verify error message is displayed on screen",
    "expected": "System displays validation error 'Invalid year format. Please use the required format.' Save action is blocked. The year format field is highlighted in red.",
    "type": "Negative",
    "priority": "High",
    "related_tc": "None"
  }}
]

Generate ALL required test cases now. Return ONLY the JSON array.
"""

    # Attempt 1
    raw    = call_llm(prompt, api_key, provider_config, max_tokens=8000)
    result = parse_json_response(raw)
    if result:
        return result

    # Attempt 2 — simpler prompt
    simple_prompt = f"""Generate test cases as JSON array for this feature.

Feature: {combined_text[:500]}

Return ONLY JSON, starting with [ and ending with ].
Each object must have exactly these 8 keys:
id, title, prerequisites, steps, expected, type, priority, related_tc

- prerequisites: "User: [role] logged in, [module] accessible, [test data] available"
- steps: "1. [action]\\n2. [action]\\n3. [action]\\n4. [action]\\n5. Verify [outcome]"
- expected: exact visible UI outcome with screen name and message text
- type: Positive or Negative or Edge
- priority: High or Medium or Low
- related_tc: None

TC IDs start from TC-{start_id:04d}.
{count_rule}
{type_rule}

[
"""
    raw    = call_llm(simple_prompt, api_key, provider_config, max_tokens=8000)
    result = parse_json_response("[" + raw if not raw.strip().startswith("[") else raw)
    if result:
        return result

    return []


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ══════════════════════════════════════════════════════════════════════════════
def build_excel(test_cases: list, feature_name: str) -> bytes:
    wb     = openpyxl.Workbook()
    h_fill = PatternFill("solid", fgColor="1A1A2E")
    h_font = Font(bold=True, color="E8F4FD", name="Calibri", size=11)
    ca     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    la     = Alignment(horizontal="left",   vertical="top",   wrap_text=True)
    thin   = Side(style="thin", color="CCCCCC")
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

    pos_count  = sum(1 for t in test_cases if "positive" in str(t.get("type","")).lower())
    neg_count  = sum(1 for t in test_cases if "negative" in str(t.get("type","")).lower())
    edge_count = sum(1 for t in test_cases if "edge"     in str(t.get("type","")).lower())

    # ── SUMMARY SHEET ──────────────────────────────────────────────────────────
    ws_s       = wb.active
    ws_s.title = "SUMMARY"
    s_hdrs = ["Feature / Module Name","Total Test Cases","Positive Count",
               "Negative Count","Edge Case Count","Acceptance Criteria Covered","Gaps Identified"]
    for c, h in enumerate(s_hdrs, 1):
        cell = ws_s.cell(row=1, column=c, value=h)
        cell.fill = h_fill; cell.font = h_font; cell.alignment = ca
        ws_s.column_dimensions[get_column_letter(c)].width = 22
    ws_s.cell(row=2, column=1, value=feature_name)
    ws_s.cell(row=2, column=2, value=len(test_cases))
    ws_s.cell(row=2, column=3, value=pos_count)
    ws_s.cell(row=2, column=4, value=neg_count)
    ws_s.cell(row=2, column=5, value=edge_count)
    ws_s.cell(row=2, column=6, value="Extracted from feature input")
    ws_s.cell(row=2, column=7, value="Legacy system integration, Performance testing")
    ws_s.freeze_panes = "A2"

    # ── TEST CASE SHEET ────────────────────────────────────────────────────────
    # FIX: Sanitise feature_name before using it as an Excel sheet name.
    # Excel rejects sheet names containing: : \ / ? * [ ]
    # and names longer than 31 characters.
    sheet_title = safe_sheet_name(feature_name, fallback="TestCases")
    ws = wb.create_sheet(title=sheet_title)

    tc_hdrs = ["Test Case ID","Test Case Title","Prerequisites",
               "Test Steps","Expected Result","Test Type","Priority","Related TC ID"]
    widths  = [14, 35, 32, 55, 42, 12, 10, 14]

    fills = {
        "pos"      : PatternFill("solid", fgColor="E2EFDA"),
        "pos_alt"  : PatternFill("solid", fgColor="D0E8C5"),
        "neg"      : PatternFill("solid", fgColor="FCE4D6"),
        "neg_alt"  : PatternFill("solid", fgColor="F0D0B8"),
        "edge"     : PatternFill("solid", fgColor="FFF2CC"),
        "edge_alt" : PatternFill("solid", fgColor="EFE8AA"),
    }

    for c, (h, w) in enumerate(zip(tc_hdrs, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = h_fill; cell.font = h_font
        cell.alignment = ca; cell.border = bdr
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for ri, tc in enumerate(test_cases, 2):
        tc_type = str(tc.get("type","")).lower()
        alt     = ri % 2 == 0
        if "negative" in tc_type:
            fill = fills["neg_alt"] if alt else fills["neg"]
        elif "edge" in tc_type:
            fill = fills["edge_alt"] if alt else fills["edge"]
        else:
            fill = fills["pos_alt"] if alt else fills["pos"]

        row_vals = [
            tc.get("id",           f"TC-{3000+ri:04d}"),
            tc.get("title",        ""),
            tc.get("prerequisites",""),
            tc.get("steps",        ""),
            tc.get("expected",     ""),
            tc.get("type",         ""),
            tc.get("priority",     ""),
            tc.get("related_tc",   "None"),
        ]
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.border = bdr; cell.fill = fill
            cell.alignment = ca if ci in [1, 6, 7, 8] else la
        ws.row_dimensions[ri].height = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# UI LAYOUT
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="app-header">
    <div class="app-title">QA Test Case Generator</div>
    <div class="app-sub">Copilot-quality enterprise test cases — switch LLM when rate limit hits</div>
    <div style="margin-top:0.8rem">
        <span class="tag">GROQ AI</span>
        <span class="tag">TOGETHER AI</span>
        <span class="tag">RAG</span>
        <span class="tag">NO LIMIT</span>
        <span class="tag">EXCEL EXPORT</span>
    </div>
</div>
""", unsafe_allow_html=True)


# ── SIDEBAR ───────────────────────────────────────────────────────────────────
with st.sidebar:

    st.markdown('<div class="sidebar-section">', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-title">🤖 Select AI Model</div>', unsafe_allow_html=True)
    st.caption("Switch to another model if you hit the rate limit")

    selected_llm = st.selectbox(
        "AI Model",
        options   = list(LLM_PROVIDERS.keys()),
        index     = 0,
        label_visibility = "hidden"
    )
    provider_cfg = LLM_PROVIDERS[selected_llm]
    limit_text   = provider_cfg["limit"]
    key_link     = provider_cfg["key_link"]

    st.markdown(f"""
    <div class="info-box">
        📊 Limit: <strong>{limit_text}</strong><br>
        🔑 Get key: <a href="{key_link}" target="_blank">{key_link.split('/')[2]}</a>
    </div>
    """, unsafe_allow_html=True)

    api_key = st.text_input(
        "API Key",
        type        = "password",
        placeholder = provider_cfg["key_hint"],
        label_visibility = "hidden"
    )
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-section">', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-title">📊 Existing Test Cases (RAG)</div>', unsafe_allow_html=True)
    st.caption("Upload master Excel → avoids duplicates, matches your company style")

    past_excel = st.file_uploader(
        "Upload existing test cases",
        type = ["xlsx", "xls"],
        label_visibility = "hidden"
    )
    if past_excel and not st.session_state.rag_loaded:
        with st.spinner("Loading all test cases into RAG..."):
            count, last_id = load_excel_to_rag(past_excel.read())
        st.success(f"✅ {count} test cases loaded | Next ID: TC-{last_id+1:04d}")
    elif past_excel and st.session_state.rag_loaded:
        st.markdown(
            f'<div class="rag-ok">✅ {st.session_state.rag_count} test cases in RAG<br>'
            f'Next TC ID: TC-{st.session_state.last_tc_id+1:04d}</div>',
            unsafe_allow_html=True
        )
    else:
        st.markdown(
            '<div class="rag-warn">⚠️ Upload your master Excel for better quality and duplicate prevention</div>',
            unsafe_allow_html=True
        )

    if st.session_state.rag_loaded:
        if st.button("🗑️ Clear RAG & Upload New"):
            st.session_state.rag_loaded = False
            st.session_state.rag_count  = 0
            st.session_state.rag_docs   = []
            st.session_state.last_tc_id = 3000
            get_vector_store().clear()
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-section">', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-title">🎛️ Settings</div>', unsafe_allow_html=True)

    auto_mode = st.toggle(
        "🤖 Auto — generate ALL scenarios",
        value   = False,
        help    = "AI decides how many test cases are needed for complete coverage"
    )
    if auto_mode:
        num_cases = 999
        st.markdown(
            '<div class="info-box">AI will generate as many test cases as needed to cover ALL scenarios (typically 40-80)</div>',
            unsafe_allow_html=True
        )
    else:
        num_cases = st.number_input(
            "Number of test cases",
            min_value = 5,
            max_value = 500,
            value     = 15,
            step      = 5,
            help      = "You can go up to 500. For complete coverage use Auto mode."
        )

    include_neg  = st.toggle("Include negative tests", value=True)
    include_edge = st.toggle("Include edge cases",      value=True)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<div class="sidebar-title">📋 Pipeline Status</div>', unsafe_allow_html=True)
    stage = st.session_state.stage or 0
    steps = [
        "🔍 Extract Text",
        "🧠 Parse Feature",
        "📚 RAG Search",
        "⚙️ Generate JSON",
        "📊 Build Excel",
    ]
    for i, label in enumerate(steps):
        color = "#2E7D32" if i < stage else ("#1565C0" if i == stage and stage > 0 else "#9B8EA0")
        icon  = "✅"      if i < stage else ("⏳"      if i == stage and stage > 0 else "○")
        st.markdown(
            f'<div style="color:{color};font-size:0.85rem;padding:0.2rem 0;">{icon} {label}</div>',
            unsafe_allow_html=True
        )


# ── MAIN LAYOUT ───────────────────────────────────────────────────────────────
col_left, col_right = st.columns([1, 1], gap="large")

with col_left:
    st.markdown('<div class="card-label">📤 Feature Input</div>', unsafe_allow_html=True)

    tab_img, tab_doc, tab_txt = st.tabs(["🖼️ Images", "📄 Documents", "📝 Text / Jira Story"])

    with tab_img:
        st.markdown(
            '<div class="info-box">💡 If OCR fails on your image, also paste the description in the Text tab for best results.</div>',
            unsafe_allow_html=True
        )
        images = st.file_uploader(
            "Upload feature images",
            type             = ["jpg","jpeg","png","webp","bmp"],
            accept_multiple_files = True,
            label_visibility = "hidden"
        )
        if images:
            for img in images:
                st.image(img, caption=img.name, use_container_width=True)

    with tab_doc:
        st.markdown(
            '<div class="info-box">💡 Supports .docx and .txt files. For VD/Figma links — open the link, copy the text content, and paste in the Text tab.</div>',
            unsafe_allow_html=True
        )
        doc_files = st.file_uploader(
            "Upload .docx or .txt files",
            type             = ["docx","txt"],
            accept_multiple_files = True,
            label_visibility = "hidden"
        )

    with tab_txt:
        manual_text = st.text_area(
            "Feature description",
            height      = 240,
            placeholder = """Paste your Jira story, acceptance criteria, or feature description.

Example:
Feature: Year Format Changes in Templates
As ignio admin, I want all templates to display dates in new year format.

Acceptance Criteria:
1. Incident report template shows date in new format
2. Email notification template uses new format
3. CSV and Excel exports maintain new format
4. Dashboard filters display new format
5. System rejects invalid year formats with error

Modules: incident report, dashboard, email notifications,
CSV export, Excel export, PDF export, alert template, rule mining

User Roles: ignio admin, regular user""",
            label_visibility = "hidden"
        )

    st.markdown("**🔗 VD / Figma / Confluence link:**")
    vd_link = st.text_input("VD Link", placeholder="https://...", label_visibility="hidden")
    if vd_link:
        st.markdown(
            '<div class="warn-box">⚠️ VD links require login — open the link, copy the text content (Ctrl+A → Ctrl+C), and paste it in the Text tab above.</div>',
            unsafe_allow_html=True
        )

    st.markdown("")

    has_image = bool(images)
    has_doc   = bool(doc_files)
    has_text  = bool(manual_text and manual_text.strip())
    has_input = has_image or has_doc or has_text
    can_run   = has_input and bool(api_key)

    if not api_key:     st.caption("🔑 Add your API key in the sidebar")
    elif not has_input: st.caption("⬆️ Upload images, documents, or paste text above")

    # ── GENERATE ──────────────────────────────────────────────────────────────
    if st.button("⚡ GENERATE TEST CASES", disabled=not can_run):

        combined_text = ""
        ocr_warnings  = []

        st.session_state.stage = 1

        if has_image:
            with st.spinner("🔍 Extracting text from images..."):
                for img_file in images:
                    text, warns = run_ocr(Image.open(img_file))
                    ocr_warnings.extend(warns)
                    if text:
                        combined_text += f"\n\n[From image: {img_file.name}]\n{text}"
                    else:
                        ocr_warnings.append(
                            f"⚠️ No text extracted from '{img_file.name}'. Please paste content in the Text tab."
                        )

        if has_doc:
            with st.spinner("📄 Reading documents..."):
                for df in doc_files:
                    raw_bytes = df.read()
                    text = extract_docx_text(raw_bytes) if df.name.endswith(".docx") \
                           else extract_txt_text(raw_bytes)
                    if text:
                        combined_text += f"\n\n[From: {df.name}]\n{text}"
                    else:
                        ocr_warnings.append(
                            f"⚠️ Could not extract text from '{df.name}'. "
                            "If it contains only images, paste the text manually."
                        )

        if has_text:
            combined_text += f"\n\n[Feature Description]\n{manual_text.strip()}"

        combined_text = combined_text.strip()
        st.session_state.ocr_warnings = ocr_warnings

        if not combined_text:
            st.error(
                "❌ No text could be extracted.\n\n"
                "**What to do:** Open your VD link/document, copy the text, "
                "and paste it in the 📝 Text tab."
            )
            st.stop()

        if len(combined_text) < 50:
            st.warning("⚠️ Very little text extracted. Consider adding more detail in the Text tab.")

        st.session_state.extracted_text = combined_text

        st.session_state.stage = 2
        with st.spinner("🧠 Parsing feature and extracting acceptance criteria..."):
            fu_prompt = f"""You are a Senior QA Architect.
Parse this feature input and produce a FEATURE UNDERSTANDING.

Return EXACTLY in this format:
Feature Name        : [name]
User Roles          : [all roles]
Screens / Pages     : [all screens and pages]
UI Components       : [all buttons, fields, dropdowns, tabs]
User Flows          : [numbered list of all user flows]
Acceptance Criteria : [numbered list — one AC per line]
Business Rules      : [constraints, validations, permissions]
Error / Edge States : [error messages, boundary conditions]
Modules Affected    : [list all specific modules and templates]

Feature Input:
{combined_text}
"""
            feature_understand = call_llm(fu_prompt, api_key, provider_cfg, max_tokens=1000)
            st.session_state.feature_understand = feature_understand

        st.session_state.stage = 3
        with st.spinner("📚 Retrieving similar past test cases..."):
            similar = rag_retrieve(feature_understand)

        st.session_state.stage = 4
        last_tc_id = st.session_state.last_tc_id or 3000
        spin_msg = (
            "⚙️ Generating ALL scenarios (auto mode — this may take 60 seconds)..."
            if auto_mode else
            f"⚙️ Generating {num_cases} test cases..."
        )
        with st.spinner(spin_msg):
            test_cases = generate_test_cases(
                combined_text      = combined_text,
                feature_understand = feature_understand,
                similar_cases      = similar,
                num_cases          = num_cases,
                auto_mode          = auto_mode,
                last_tc_id         = last_tc_id,
                api_key            = api_key,
                provider_config    = provider_cfg,
                include_neg        = include_neg,
                include_edge       = include_edge,
            )

            if not test_cases:
                st.error(
                    "❌ Could not generate test cases.\n\n"
                    "**Try these:**\n"
                    "1. Add more detail to your feature description\n"
                    "2. Switch to a different AI model in the sidebar\n"
                    "3. Wait 30 seconds (rate limit) and try again\n"
                    "4. Reduce the number of test cases"
                )
                st.stop()

            st.session_state.test_cases_parsed = test_cases
            st.session_state.tc_count          = len(test_cases)
            st.session_state.pos_count         = sum(1 for t in test_cases if "positive" in t.get("type","").lower())
            st.session_state.neg_count         = sum(1 for t in test_cases if "negative" in t.get("type","").lower())
            st.session_state.edge_count        = sum(1 for t in test_cases if "edge"     in t.get("type","").lower())
            st.session_state.llm_used          = selected_llm

        st.session_state.stage = 5
        with st.spinner("📊 Building formatted Excel with color coding..."):
            # FIX: Determine feature_name then sanitise it via safe_sheet_name
            # before passing to build_excel so the sheet title is always valid.
            feature_name = "Feature"
            if images:
                feature_name = os.path.splitext(images[0].name)[0]
            elif doc_files:
                feature_name = os.path.splitext(doc_files[0].name)[0]
            elif manual_text:
                first        = manual_text.strip().split("\n")[0]
                feature_name = first[:50].replace("Feature:","").strip() or "Feature"

            excel = build_excel(test_cases, feature_name)
            st.session_state.excel_bytes = excel

        st.rerun()


# ── RESULTS ───────────────────────────────────────────────────────────────────
with col_right:
    st.markdown('<div class="card-label">📋 Results</div>', unsafe_allow_html=True)

    for w in (st.session_state.ocr_warnings or []):
        st.markdown(f'<div class="warn-box">{w}</div>', unsafe_allow_html=True)

    if not st.session_state.test_cases_parsed:
        st.markdown("""
        <div style="background:#FFFFFF;border:1px solid #E8E5DF;border-radius:12px;
                    padding:4rem 2rem;text-align:center;">
            <div style="font-size:3rem">⏳</div>
            <div style="margin-top:0.8rem;font-size:0.9rem;color:#9B8EA0;">
                Results will appear here after generation
            </div>
        </div>""", unsafe_allow_html=True)
    else:
        tc   = st.session_state.tc_count   or 0
        pos  = st.session_state.pos_count  or 0
        neg  = st.session_state.neg_count  or 0
        edge = st.session_state.edge_count or 0
        rag  = st.session_state.rag_count  or 0
        llm  = st.session_state.llm_used   or ""

        if llm:
            st.markdown(f'<span class="llm-badge">Generated by: {llm}</span>', unsafe_allow_html=True)

        st.markdown(
            f'<div class="success-box">✅ &nbsp; <strong>{tc} enterprise test cases generated!</strong></div>',
            unsafe_allow_html=True
        )
        st.markdown(f"""
        <div class="metric-strip">
            <div class="metric"><div class="metric-val" style="color:#2E7D32">{pos}</div><div class="metric-lbl">✅ Positive</div></div>
            <div class="metric"><div class="metric-val" style="color:#C00000">{neg}</div><div class="metric-lbl">❌ Negative</div></div>
            <div class="metric"><div class="metric-val" style="color:#F59E0B">{edge}</div><div class="metric-lbl">⚠️ Edge</div></div>
            <div class="metric"><div class="metric-val">{rag}</div><div class="metric-lbl">📚 RAG</div></div>
        </div>""", unsafe_allow_html=True)

        t1, t2, t3 = st.tabs(["📊 Test Cases Preview", "🧠 Feature Understanding", "🔍 Extracted Text"])

        with t1:
            tcs     = st.session_state.test_cases_parsed
            preview = "\n\n".join(
                f"{t.get('id','')} | {t.get('type','')} | {t.get('priority','')}\n"
                f"Title: {t.get('title','')}\n"
                f"Prerequisites: {t.get('prerequisites','')}\n"
                f"Steps:\n{t.get('steps','')}\n"
                f"Expected: {t.get('expected','')}"
                for t in tcs[:5]
            )
            if len(tcs) > 5:
                preview += f"\n\n... and {len(tcs)-5} more test cases in the Excel file"
            st.markdown(f'<div class="result-preview">{preview}</div>', unsafe_allow_html=True)

        with t2:
            st.markdown(
                f'<div class="result-preview">{st.session_state.feature_understand or "—"}</div>',
                unsafe_allow_html=True
            )

        with t3:
            st.markdown(
                f'<div class="result-preview">{st.session_state.extracted_text or "—"}</div>',
                unsafe_allow_html=True
            )

        st.markdown("---")

        fname = "Feature"
        if st.session_state.test_cases_parsed:
            fname = st.session_state.feature_understand.split("\n")[0].replace("Feature Name :", "").strip()[:30] or "Feature"
        # Sanitise download filename too
        fname = safe_sheet_name(fname, fallback="Feature")

        st.download_button(
            label     = "⬇️ DOWNLOAD EXCEL TEST CASES",
            data      = st.session_state.excel_bytes,
            file_name = f"{fname}_TestCases.xlsx",
            mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        if st.button("🔄 Generate New"):
            for k in ["extracted_text","feature_understand","test_cases_parsed",
                      "excel_bytes","tc_count","pos_count","neg_count","edge_count",
                      "ocr_warnings","llm_used"]:
                st.session_state[k] = [] if k in ["test_cases_parsed","ocr_warnings"] else None
            st.session_state.stage = 0
            st.rerun()
